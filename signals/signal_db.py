"""
Signal History Database (SQLite)
==================================
Efficient storage and retrieval of signal history.
Replacement for Excel-based storage.

Features:
- Batch insert (single connection)
- Duplicate detection
- Fast queries with indexes
- Export to Excel/CSV
- VPS-friendly (no file lock issues)

Usage:
    from signal_db import SignalDatabase
    
    db = SignalDatabase()
    db.insert_signal(signal_data)
    db.get_all_signals()
    db.export_to_excel("report.xlsx")
"""

import sqlite3
import os
import logging
from datetime import datetime
from typing import Dict, List, Optional, Tuple
from contextlib import contextmanager

logger = logging.getLogger(__name__)


class SignalDatabase:
    """
    SQLite database untuk signal history.
    Efficient, VPS-friendly, no file lock issues.
    """
    
    def __init__(self, db_path: str = "data/signals.db"):
        self.db_path = db_path
        
        # Ensure directory exists
        os.makedirs(os.path.dirname(db_path), exist_ok=True)
        
        # Initialize database
        self._create_tables()
        logger.info(f"✅ Signal database initialized: {db_path}")
    
    @contextmanager
    def get_connection(self, autocommit: bool = False):
        """Get database connection with proper settings"""
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")  # Better concurrency
        conn.execute("PRAGMA synchronous=NORMAL")  # Faster writes
        conn.execute("PRAGMA cache_size=10000")  # Larger cache
        
        try:
            yield conn
            if autocommit:
                conn.commit()
        except Exception as e:
            conn.rollback()
            raise e
        finally:
            conn.close()
    
    def _create_tables(self):
        """Create tables and indexes if not exist"""
        with self.get_connection(autocommit=True) as conn:
            cursor = conn.cursor()
            
            # Signals table
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS signals (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    symbol TEXT NOT NULL,
                    price REAL NOT NULL,
                    recommendation TEXT NOT NULL,
                    rsi TEXT,
                    macd TEXT,
                    ma_trend TEXT,
                    bollinger TEXT,
                    volume TEXT,
                    ml_confidence REAL,
                    combined_strength REAL,
                    analysis TEXT,
                    signal_time TEXT,
                    received_at TEXT NOT NULL,
                    received_date TEXT NOT NULL,
                    source TEXT DEFAULT 'telegram',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            # Create indexes for fast queries
            cursor.execute('''
                CREATE INDEX IF NOT EXISTS idx_signals_symbol 
                ON signals(symbol)
            ''')
            cursor.execute('''
                CREATE INDEX IF NOT EXISTS idx_signals_received_date 
                ON signals(received_date)
            ''')
            cursor.execute('''
                CREATE INDEX IF NOT EXISTS idx_signals_recommendation 
                ON signals(recommendation)
            ''')
            cursor.execute('''
                CREATE INDEX IF NOT EXISTS idx_signals_received_at 
                ON signals(received_at)
            ''')
            
            # Metadata table for tracking
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS signal_metadata (
                    key TEXT PRIMARY KEY,
                    value TEXT,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            logger.debug("📊 Database tables and indexes created")
    
    def insert_signal(self, signal: Dict, received_at: datetime, retries: int = 3) -> int:
        """
        Insert single signal with retry on database lock
        Returns signal ID, or -1 if duplicate, or -2 if all retries failed
        """
        for attempt in range(retries):
            try:
                with self.get_connection(autocommit=True) as conn:
                    cursor = conn.cursor()

                    # Check duplicate
                    dup_key = (
                        received_at.strftime("%Y-%m-%d"),
                        signal["symbol"].upper(),
                        signal["rec"].upper(),
                        signal["price"]
                    )

                    if self._is_duplicate(cursor, dup_key):
                        logger.debug(f"⏭️ Duplicate signal skipped: {signal['symbol']} @ {received_at.strftime('%H:%M')}")
                        return -1

                    cursor.execute('''
                        INSERT INTO signals (
                            symbol, price, recommendation, rsi, macd, ma_trend,
                            bollinger, volume, ml_confidence, combined_strength,
                            analysis, signal_time, received_at, received_date
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        signal["symbol"].upper(),
                        self._parse_float(signal["price"]),
                        signal["rec"].upper(),
                        signal.get("rsi", "—"),
                        signal.get("macd", "—"),
                        signal.get("ma", "—"),
                        signal.get("bollinger", "—"),
                        signal.get("volume", "—"),
                        self._parse_confidence(signal.get("confidence", "0%")),
                        self._parse_float(signal.get("strength", "0")),
                        signal.get("analysis", ""),
                        signal.get("signal_time", received_at.strftime("%H:%M:%S")),
                        received_at.strftime("%Y-%m-%d %H:%M:%S"),
                        received_at.strftime("%Y-%m-%d")
                    ))

                    return cursor.lastrowid

            except sqlite3.OperationalError as e:
                if "locked" in str(e).lower() and attempt < retries - 1:
                    logger.warning(f"🔒 Signal DB locked, retry {attempt + 1}/{retries}")
                    import time
                    time.sleep(0.5)  # Wait 500ms before retry
                else:
                    logger.error(f"❌ Signal insert failed after {attempt + 1} attempts: {e}")
                    return -2

            except Exception as e:
                logger.error(f"❌ Signal insert error: {e}")
                return -2
    
    def insert_signals_batch(self, signals: List[Tuple[Dict, datetime]], retries: int = 3) -> int:
        """
        Batch insert multiple signals with retry on database lock
        Returns count of inserted signals
        """
        if not signals:
            return 0

        inserted = 0
        skipped = 0

        for attempt in range(retries):
            try:
                with self.get_connection() as conn:
                    cursor = conn.cursor()

                    for signal, received_at in signals:
                        try:
                            # Check duplicate
                            dup_key = (
                                received_at.strftime("%Y-%m-%d"),
                                signal["symbol"].upper(),
                                signal["rec"].upper(),
                                signal["price"]
                            )

                            if self._is_duplicate(cursor, dup_key):
                                skipped += 1
                                continue

                            cursor.execute('''
                                INSERT INTO signals (
                                    symbol, price, recommendation, rsi, macd, ma_trend,
                                    bollinger, volume, ml_confidence, combined_strength,
                                    analysis, signal_time, received_at, received_date
                                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            ''', (
                                signal["symbol"].upper(),
                                self._parse_float(signal["price"]),
                                signal["rec"].upper(),
                                signal.get("rsi", "—"),
                                signal.get("macd", "—"),
                                signal.get("ma", "—"),
                                signal.get("bollinger", "—"),
                                signal.get("volume", "—"),
                                self._parse_confidence(signal.get("confidence", "0%")),
                                self._parse_float(signal.get("strength", "0")),
                                signal.get("analysis", ""),
                                signal.get("signal_time", received_at.strftime("%H:%M:%S")),
                                received_at.strftime("%Y-%m-%d %H:%M:%S"),
                                received_at.strftime("%Y-%m-%d")
                            ))

                            inserted += 1

                        except Exception as e:
                            logger.error(f"❌ Failed to insert signal: {e}")
                            skipped += 1
                            continue

                    # Commit all at once
                    conn.commit()
                    logger.info(f"✅ Batch insert: {inserted} inserted, {skipped} skipped")
                    return inserted

            except sqlite3.OperationalError as e:
                if "locked" in str(e).lower() and attempt < retries - 1:
                    logger.warning(f"🔒 Signal DB locked (batch), retry {attempt + 1}/{retries}")
                    import time
                    time.sleep(0.5)
                else:
                    logger.error(f"❌ Batch insert failed after {attempt + 1} attempts: {e}")
                    return 0

            except Exception as e:
                logger.error(f"❌ Batch insert error: {e}")
                return 0

        return 0
    
    def _is_duplicate(self, cursor, dup_key: Tuple) -> bool:
        """Check if signal already exists"""
        cursor.execute('''
            SELECT COUNT(*) as cnt FROM signals
            WHERE received_date = ?
            AND symbol = ?
            AND recommendation = ?
            AND price = ?
        ''', dup_key)
        
        result = cursor.fetchone()
        return result['cnt'] > 0 if result else False
    
    def get_signals(
        self,
        symbol: Optional[str] = None,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        recommendation: Optional[str] = None,
        limit: int = 100,
        order: str = "DESC"
    ) -> List[Dict]:
        """
        Query signals with filters
        
        Args:
            symbol: Filter by symbol (e.g., "BTCIDR")
            start_date: Filter from date (YYYY-MM-DD)
            end_date: Filter to date (YYYY-MM-DD)
            recommendation: Filter by rec (BUY/SELL/HOLD)
            limit: Max results
            order: ASC or DESC
        
        Returns:
            List of signal dicts
        """
        query = "SELECT * FROM signals WHERE 1=1"
        params = []
        
        if symbol:
            query += " AND symbol = ?"
            params.append(symbol.upper())
        
        if start_date:
            query += " AND received_date >= ?"
            params.append(start_date)
        
        if end_date:
            query += " AND received_date <= ?"
            params.append(end_date)
        
        if recommendation:
            query += " AND recommendation = ?"
            params.append(recommendation.upper())
        
        query += f" ORDER BY received_at {order} LIMIT ?"
        params.append(limit)
        
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(query, params)
            rows = cursor.fetchall()
            
            return [dict(row) for row in rows]
    
    def cleanup_incomplete_signals(self) -> int:
        """
        Delete signals with incomplete data (rsi='—', macd='—', etc.)
        Checks for both em dash (—) and triple hyphens (---)
        Returns number of deleted rows.
        """
        with self.get_connection(autocommit=True) as conn:
            cursor = conn.cursor()

            # Count incomplete signals first
            # Check for BOTH em dash (—) and triple hyphens (---)
            cursor.execute('''
                SELECT COUNT(*) FROM signals
                WHERE rsi IN ('—', '---') OR macd IN ('—', '---')
                OR ma_trend IN ('—', '---') OR bollinger IN ('—', '---')
                OR volume IN ('—', '---')
            ''')
            incomplete_count = cursor.fetchone()[0]

            if incomplete_count > 0:
                # Delete incomplete signals
                cursor.execute('''
                    DELETE FROM signals
                    WHERE rsi IN ('—', '---') OR macd IN ('—', '---')
                    OR ma_trend IN ('—', '---') OR bollinger IN ('—', '---')
                    OR volume IN ('—', '---')
                ''')
                deleted = cursor.rowcount
                logger.info(f"🗑️ Cleaned up {deleted} incomplete signal records")
                return deleted
            else:
                logger.info("✅ No incomplete signals found")
                return 0

    def get_total_count(self) -> int:
        """Get total number of signals in database"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT COUNT(*) FROM signals')
            return cursor.fetchone()[0]

    def get_stats(self) -> Dict:
        """Get signal statistics"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            
            # Total signals
            cursor.execute("SELECT COUNT(*) as total FROM signals")
            total = cursor.fetchone()['total']
            
            # By recommendation
            cursor.execute('''
                SELECT recommendation, COUNT(*) as count
                FROM signals
                GROUP BY recommendation
                ORDER BY count DESC
            ''')
            by_rec = {row['recommendation']: row['count'] for row in cursor.fetchall()}
            
            # By symbol (top 10)
            cursor.execute('''
                SELECT symbol, COUNT(*) as count
                FROM signals
                GROUP BY symbol
                ORDER BY count DESC
                LIMIT 10
            ''')
            by_symbol = {row['symbol']: row['count'] for row in cursor.fetchall()}
            
            # Date range
            cursor.execute('''
                SELECT MIN(received_date) as first, MAX(received_date) as last
                FROM signals
            ''')
            date_range = cursor.fetchone()
            
            # Average confidence
            cursor.execute('''
                SELECT AVG(ml_confidence) as avg_conf
                FROM signals
                WHERE ml_confidence > 0
            ''')
            avg_conf = cursor.fetchone()['avg_conf'] or 0
            
            return {
                "total_signals": total,
                "by_recommendation": by_rec,
                "top_symbols": by_symbol,
                "date_range": {
                    "first": date_range['first'],
                    "last": date_range['last']
                },
                "avg_confidence": avg_conf
            }
    
    def get_recent_signals(self, limit: int = 10) -> List[Dict]:
        """Get most recent signals"""
        return self.get_signals(limit=limit, order="DESC")
    
    def delete_old_signals(self, days: int = 90) -> int:
        """Delete signals older than N days"""
        from datetime import timedelta
        cutoff = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")
        
        with self.get_connection(autocommit=True) as conn:
            cursor = conn.cursor()
            cursor.execute('DELETE FROM signals WHERE received_date < ?', (cutoff,))
            deleted = cursor.rowcount
            
            if deleted > 0:
                logger.info(f"🗑️ Deleted {deleted} old signals (older than {days} days)")
                conn.execute('VACUUM')  # Reclaim space
            
            return deleted
    
    def export_to_excel(self, filepath: str, filters: Dict = None) -> int:
        """
        Export signals to Excel file
        Returns row count
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Get signals
            signals = self.get_signals(**(filters or {}))
            
            if not signals:
                logger.warning("No signals to export")
                return 0
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Signal History"
            ws.freeze_panes = "A2"
            
            # Headers
            headers = [
                "No", "Date", "Time", "Symbol", "Price (IDR)",
                "Recommendation", "RSI (14)", "MACD", "MA Trend",
                "Bollinger", "Volume", "ML Confidence",
                "Combined Strength", "Analysis", "Received At"
            ]
            
            col_widths = [5, 12, 10, 14, 16, 14, 12, 12, 12, 12, 10, 14, 16, 40, 16]
            
            # Style
            header_fill = PatternFill("solid", start_color="1F4E79")
            header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
            data_font = Font(name="Arial", size=10)
            alt_fill = PatternFill("solid", start_color="D6E4F0")
            green_font = Font(name="Arial", size=10, color="1A7A1A", bold=True)
            red_font = Font(name="Arial", size=10, color="CC0000", bold=True)
            center = Alignment(horizontal="center", vertical="center")
            left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )
            
            # Write headers
            for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
                cell.border = thin_border
                ws.column_dimensions[get_column_letter(col_idx)].width = width
            
            # Write data
            for idx, signal in enumerate(signals, 1):
                row = [
                    idx,
                    signal['received_date'],
                    signal['signal_time'],
                    signal['symbol'],
                    signal['price'],
                    signal['recommendation'],
                    signal['rsi'],
                    signal['macd'],
                    signal['ma_trend'],
                    signal['bollinger'],
                    signal['volume'],
                    f"{signal['ml_confidence']:.1%}" if signal['ml_confidence'] else "—",
                    f"{signal['combined_strength']:.2f}" if signal['combined_strength'] else "—",
                    signal['analysis'],
                    signal['received_at']
                ]
                
                use_alt = (idx % 2 == 0)
                
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=idx + 1, column=col_idx, value=value)
                    cell.font = data_font
                    cell.border = thin_border
                    cell.alignment = left_wrap if col_idx in (4, 14) else center
                    if use_alt:
                        cell.fill = alt_fill
                
                # Color recommendation
                rec_cell = ws.cell(row=idx + 1, column=6)
                if signal['recommendation'] in ('BUY', 'STRONG_BUY'):
                    rec_cell.font = green_font
                elif signal['recommendation'] in ('SELL', 'STRONG_SELL'):
                    rec_cell.font = red_font
                
                ws.row_dimensions[idx + 1].height = 18
            
            wb.save(filepath)
            logger.info(f"📄 Exported {len(signals)} signals to: {filepath}")
            return len(signals)
            
        except ImportError:
            logger.error("❌ openpyxl not installed. Run: pip install openpyxl")
            return 0
    
    def vacuum(self):
        """Optimize database size"""
        with self.get_connection(autocommit=True) as conn:
            conn.execute('VACUUM')
            logger.info("🗜️ Database vacuumed")
    
    def _parse_float(self, value) -> float:
        """Parse value to float"""
        if isinstance(value, (int, float)):
            return float(value)
        try:
            return float(str(value).replace(",", "").strip())
        except Exception as e:
            logger.debug(f"Failed to parse float from '{value}': {e}")
            return 0.0

    def _parse_confidence(self, value) -> float:
        """Parse confidence to 0-1 range"""
        if isinstance(value, (int, float)):
            return value / 100 if value > 1 else value
        try:
            val = float(str(value).replace("%", "").strip())
            return val / 100 if val > 1 else val
        except Exception as e:
            logger.debug(f"Failed to parse confidence from '{value}': {e}")
            return 0.0
